# REAL-TIME SOCCER ANALYTICS

# Features:
# - Roboflow 4-class detection (ball, goalkeeper, player, referee)
# - SiGLIP-based team classification (TeamClassifier)
# - ByteTrack player tracking with persistent IDs
# - Field keypoints + homography (smoothed)
# - Pitch projection with team dots + ball
# - Voronoi diagram (standard + smooth blend)
# - Rich annotators (Ellipse, Label, Triangle)
# - Live OpenCV windows

import os
import csv
import json
import math
import time
import argparse
import threading

# Allow unsupported MPS ops to fall back to CPU instead of erroring
os.environ.setdefault("PYTORCH_ENABLE_MPS_FALLBACK", "1")

import cv2
import torch
import numpy as np
from collections import deque, defaultdict
from typing import Optional

import supervision as sv
from inference import get_model

from sports.annotators.soccer import (
    draw_pitch,
    draw_points_on_pitch,
    draw_pitch_voronoi_diagram
)
from sports.common.team import TeamClassifier
from sports.common.view import ViewTransformer
from sports.configs.soccer import SoccerPitchConfiguration

from broadcast import StatsBroadcaster


# CONFIG

API_KEY = os.environ.get("ROBOFLOW_API_KEY")
CONFIG = SoccerPitchConfiguration()

# Class IDs (must match the spen-rtgs-oc4ez/4 model)
BALL_ID = 0
GOALKEEPER_ID = 1
PLAYER_ID = 2
REFEREE_ID = 3

# Max detection counts per class
MAX_BALLS = 1
MAX_GOALKEEPERS = 2
MAX_PLAYERS = 20
MAX_REFEREES = 3

# Calibration settings
CALIBRATION_FRAMES = 100
CALIBRATION_STRIDE = 3

# Homography smoothing window
MAXLEN = 5

# Refresh field-keypoint detection (which drives the homography) only every N
# processed frames and reuse the smoothed transform in between. A broadcast
# camera pans slowly, so re-detecting every frame roughly doubles inference
# cost for little benefit.
FIELD_DETECT_INTERVAL = 15

# A player's team never changes, so classify each track with the SiGLIP model
# only until it has accumulated this many votes, then lock its team and stop
# re-embedding it every frame.
TEAM_VOTE_FRAMES = 5

# Expected Goals (xG) settings
GOAL_WIDTH = 7.32          # meters, standard goal mouth
BALL_HISTORY_LEN = 10      # frames of ball positions kept for velocity
SHOT_MIN_SPEED = 8.0       # m/s; min ball speed over the window to count as a shot
SHOT_MAX_DISTANCE = 20.0   # meters; shots only count this close to a goal
SHOT_MAX_STEP = 5.0        # meters/step floor for the glitch guard (low stride)
MAX_BALL_SPEED = 35.0      # m/s; physical cap for the stride-aware glitch guard
SHOT_AIM_ANGLE = 40.0      # degrees; max angle between ball velocity and the goal
SHOT_MIN_ADVANCE = 0.5     # meters; min net x-advance toward goal to be goal-directed
SHOT_COOLDOWN_FRAMES = 150  # frames to wait before another shot can register (~5s)
# Loose pre-gate for diagnostic logging of shot *candidates* (rejected included),
# so thresholds can be re-tuned offline from one run instead of re-processing.
SHOT_LOG_DISTANCE = 30.0   # meters; log candidates within this range of a goal
SHOT_LOG_MIN_SPEED = 3.0   # m/s; and at least this fast
XG_DISPLAY_DURATION = 90   # frames to keep a shot marker on the radar

# Zone-entry tracking settings ("Key Areas in the Final Third").
# The penalty box (near goal) is the BOX zone. Beyond it, an outer band runs
# from the box's front edge to the final-third line; across the box width that
# band splits into CENTRAL (inner) and two HALF-spaces (outer strips). Anything
# outside the box width within the final third is WIDE.
FINAL_THIRD_FRAC = 1.0 / 3.0           # depth of the final third (frac of length)
BOX_DEPTH = CONFIG.penalty_box_length  # cm, penalty-box depth from the goal line
CENTRAL_WIDTH = 2700                   # cm, width of the central key zone
BOX_LO = CONFIG.width / 2 - CONFIG.penalty_box_width / 2   # box width, low y
BOX_HI = CONFIG.width / 2 + CONFIG.penalty_box_width / 2   # box width, high y
CEN_LO = CONFIG.width / 2 - CENTRAL_WIDTH / 2              # central zone, low y
CEN_HI = CONFIG.width / 2 + CENTRAL_WIDTH / 2             # central zone, high y
ZONE_NAMES = ["BOX", "CENTRAL", "HALF_L", "HALF_R", "WIDE_L", "WIDE_R"]
ZONE_LABELS = {
    "BOX": "Box", "CENTRAL": "Central", "HALF_L": "Half-sp L",
    "HALF_R": "Half-sp R", "WIDE_L": "Wide L", "WIDE_R": "Wide R",
}
# A zone entry only counts after the SAME team holds the ball in the SAME zone
# for this many consecutive frames. Debounces ball/possession jitter that would
# otherwise inflate the count with phantom transitions.
ZONE_CONFIRM_FRAMES = 8
# Attack-direction assumption: Team 0 (cyan) attacks the goal at x = length.
# Set to False if your teams attack the other way.
TEAM0_ATTACKS_RIGHT = True

# Team-classification settings. Embedding only the torso band (skipping head,
# legs, shorts and surrounding grass) concentrates the SiGLIP embedding on
# shirt colour/pattern, which is what actually separates the two teams.
JERSEY_TOP_FRAC = 0.15     # top of the torso band, as a fraction of box height
JERSEY_BOT_FRAC = 0.55     # bottom of the torso band, as a fraction of box height
# When the source is seekable (a video file), sample this many frames spread
# evenly across the whole match for a representative fit rather than only the
# opening seconds. Live streams fall back to sequential sampling.
CALIBRATION_SAMPLES = 60
# Rolling buffer of recent jersey crops kept for on-demand re-calibration ('c').
RECAL_BUFFER = 400

# Coach reporting: emit a summary every N minutes of match (video) time
INTERVAL_MINUTES = 15

# Live-stream resilience (Veo Live HLS/DASH, RTSP/RTMP): tolerate brief stalls
LIVE_MAX_MISSES = 300      # consecutive empty reads before treating as ended
LIVE_RETRY_DELAY = 0.1     # seconds to wait between retries on an empty read

# xG logistic-model coefficients: log_odds = b0 + b_dist*distance + b_angle*angle
# Defaults are calibrated to typical Wyscout open-play xG by location
# (e.g. ~0.45 at 6 m central, ~0.19 at 11 m, ~0.08 at the box edge).
# train_xg.py can overwrite them via xg_coeffs.json.
XG_COEFFS = {"intercept": -0.845, "distance": -0.131, "angle": 1.308}

_XG_COEFFS_PATH = os.path.join(os.path.dirname(__file__), "xg_coeffs.json")
if os.path.exists(_XG_COEFFS_PATH):
    try:
        with open(_XG_COEFFS_PATH) as _f:
            _loaded = json.load(_f)
        XG_COEFFS.update({k: float(_loaded[k]) for k in XG_COEFFS if k in _loaded})
        print(f"✅ Loaded trained xG coefficients from {_XG_COEFFS_PATH}")
    except (ValueError, KeyError, OSError) as _e:
        print(f"⚠️  Could not load xg_coeffs.json ({_e}); using defaults")


# LOAD MODELS

PLAYER_DETECTION_MODEL = get_model(
    model_id="spen-rtgs-oc4ez/4",
    api_key=API_KEY
)

FIELD_DETECTION_MODEL = get_model(
    model_id="football-field-detection-f07vi/14",
    api_key=API_KEY
)



# HELPERS

def resolve_goalkeepers_team_id(
    players: sv.Detections,
    goalkeepers: sv.Detections
) -> np.ndarray:
    """Assign each goalkeeper to the nearest team based on centroid distance."""
    players_xy = players.get_anchors_coordinates(sv.Position.BOTTOM_CENTER)
    if (len(players_xy[players.class_id == 0]) == 0 or
            len(players_xy[players.class_id == 1]) == 0):
        return np.zeros(len(goalkeepers), dtype=int)

    team0 = players_xy[players.class_id == 0].mean(axis=0)
    team1 = players_xy[players.class_id == 1].mean(axis=0)

    team_ids = []
    for gk_xy in goalkeepers.get_anchors_coordinates(sv.Position.BOTTOM_CENTER):
        team_ids.append(
            0 if np.linalg.norm(gk_xy - team0) < np.linalg.norm(gk_xy - team1) else 1
        )
    return np.array(team_ids, dtype=int)


def calculate_xg(shot_x: float, shot_y: float, goal_x: float,
                 config: SoccerPitchConfiguration) -> float:
    """
    Estimate Expected Goals (xG) from a shot location using a simple
    distance + angle logistic model. Coordinates are in pitch space (cm).

    shot_x, shot_y: shot location on the pitch
    goal_x: x-coordinate of the goal being attacked (0 or config.length)
    """
    # Work in meters (sports pitch config is in centimeters)
    sx, sy = shot_x / 100.0, shot_y / 100.0
    gx = goal_x / 100.0
    gy = (config.width / 100.0) / 2.0

    # Distance from shot to goal center
    distance = math.hypot(gx - sx, gy - sy)

    # Angle of the visible goal mouth from the shot location
    post1 = (gx, gy - GOAL_WIDTH / 2.0)
    post2 = (gx, gy + GOAL_WIDTH / 2.0)
    a1 = math.atan2(post1[1] - sy, post1[0] - sx)
    a2 = math.atan2(post2[1] - sy, post2[0] - sx)
    angle = abs(a2 - a1)

    # Logistic model. Coefficients come from XG_COEFFS, which may be the
    # rough defaults or values trained on real data via train_xg.py.
    log_odds = (
        XG_COEFFS["intercept"]
        + XG_COEFFS["distance"] * distance
        + XG_COEFFS["angle"] * angle
    )
    # Clamp to avoid OverflowError in math.exp for extreme log-odds values.
    log_odds = max(-60.0, min(60.0, log_odds))
    xg = 1.0 / (1.0 + math.exp(-log_odds))
    return float(min(max(xg, 0.01), 0.99))


def detect_shot(ball_history: deque, config: SoccerPitchConfiguration,
                frame_gap: int = 1, fps: float = 30.0):
    """
    Detect a shot from recent pitch-space ball positions.

    A movement is treated as a shot only when it is (a) fast enough,
    (b) heading consistently toward one goal across the window, and
    (c) close enough to that goal. This rejects ordinary passes and
    clearances that merely happen to be quick.

    `frame_gap` (the processing stride) and `fps` convert the window's net
    displacement into a physical speed (m/s), so the same speed gate applies
    regardless of stride. The aim check uses the angle between the ball's
    velocity and the line to the goal centre, rejecting fast crosses, switches
    and clearances that merely head goal-ward but are not on target.

    Returns (is_shot, velocity_x, info). `info` is a diagnostics dict with the
    computed features (speed_mps, dist_m, aim_deg, max_step_m), a `reason`
    string, and a `candidate` flag (loosely-gated, for logging rejects too).
    velocity_x sign indicates direction: positive => goal at x = config.length,
    negative => goal at x = 0.
    """
    info = {"reason": "no_data", "candidate": False}
    if len(ball_history) < 4:
        return False, 0.0, info

    gap = max(1, frame_gap)
    recent = list(ball_history)[-4:]
    window_s = 3 * gap / max(1.0, fps)   # time spanned by the 4-sample window

    # Net displacement over the window -> physical speed (m/s)
    dx_cm = recent[-1][0] - recent[0][0]
    dy_cm = recent[-1][1] - recent[0][1]
    vx = dx_cm / 100.0
    speed_mps = math.hypot(dx_cm, dy_cm) / 100.0 / window_s

    goal_x = config.length if vx >= 0 else 0.0
    dist_to_goal = abs(goal_x - recent[-1][0]) / 100.0

    # Largest single inter-frame step (meters) for the glitch guard.
    max_step_obs = max(
        math.hypot(recent[i + 1][0] - recent[i][0],
                   recent[i + 1][1] - recent[i][1]) / 100.0
        for i in range(len(recent) - 1))
    max_step = max(SHOT_MAX_STEP, MAX_BALL_SPEED * gap / max(1.0, fps))

    # Aim: angle between the velocity vector and the line to the goal centre.
    to_goal = math.atan2(config.width / 2.0 - recent[-1][1], goal_x - recent[-1][0])
    vel_dir = math.atan2(dy_cm, dx_cm)
    aim = abs((vel_dir - to_goal + math.pi) % (2 * math.pi) - math.pi)
    aim_deg = math.degrees(aim)

    info = {
        "speed_mps": round(speed_mps, 1),
        "dist_m": round(dist_to_goal, 1),
        "aim_deg": round(aim_deg, 0),
        "max_step_m": round(max_step_obs, 1),
        "reason": "ok",
        # Loose pre-gate: anything fast-ish near a goal is worth logging.
        "candidate": (dist_to_goal <= SHOT_LOG_DISTANCE
                      and speed_mps >= SHOT_LOG_MIN_SPEED),
    }

    if speed_mps < SHOT_MIN_SPEED:
        info["reason"] = "slow"
        return False, vx, info
    if max_step_obs > max_step:
        info["reason"] = "glitch"
        return False, vx, info

    step_dx = [recent[i + 1][0] - recent[i][0] for i in range(len(recent) - 1)]
    if not (all(d > 0 for d in step_dx) or all(d < 0 for d in step_dx)):
        info["reason"] = "inconsistent"
        return False, vx, info
    if dist_to_goal > SHOT_MAX_DISTANCE:
        info["reason"] = "far"
        return False, vx, info
    if abs(dx_cm) / 100.0 < SHOT_MIN_ADVANCE:
        info["reason"] = "sideways"
        return False, vx, info
    if aim_deg > SHOT_AIM_ANGLE:
        info["reason"] = "wide"
        return False, vx, info

    return True, vx, info


def jersey_crop(frame: np.ndarray, xyxy: np.ndarray) -> np.ndarray:
    """
    Crop the torso/jersey band of a player box. Skips the head and everything
    below the shirt (shorts, legs, grass) so the embedding focuses on shirt
    colour. Falls back to the full box for tiny/degenerate detections.
    """
    x1, y1, x2, y2 = (int(v) for v in xyxy)
    h = y2 - y1
    ty1 = y1 + int(JERSEY_TOP_FRAC * h)
    ty2 = y1 + int(JERSEY_BOT_FRAC * h)
    if ty2 - ty1 < 4 or x2 - x1 < 4:
        return sv.crop_image(frame, xyxy)
    return sv.crop_image(frame, np.array([x1, ty1, x2, ty2]))


def nearest_team(point_xy: np.ndarray, pitch_detections: sv.Detections) -> int:
    """Return the team id (0 or 1) of the closest player to a pitch point."""
    if len(pitch_detections) == 0:
        return 0
    players_xy = pitch_detections.get_anchors_coordinates(
        sv.Position.BOTTOM_CENTER)
    dists = np.linalg.norm(players_xy - point_xy, axis=1)
    closest = int(np.argmin(dists))
    team = int(pitch_detections.class_id[closest])
    return team if team in (0, 1) else 0


def team_attacks_right(team: int, team0_right: bool = TEAM0_ATTACKS_RIGHT) -> bool:
    """True if the given team attacks the goal at x = length.

    `team0_right` is whether Team 0 attacks the x=length goal for this run;
    pass the half-adjusted value so the side switch is handled.
    """
    return team0_right if team == 0 else not team0_right


def _dominant_color_bgr(crops: list) -> np.ndarray:
    """Median BGR colour across a set of crops (robust to outlier pixels)."""
    medians = []
    for c in crops:
        if c is None or getattr(c, "size", 0) == 0:
            continue
        medians.append(np.median(c.reshape(-1, 3), axis=0))
    if not medians:
        return np.zeros(3, dtype=float)
    return np.median(np.array(medians), axis=0)


def _luminance_bgr(bgr: np.ndarray) -> float:
    """Perceived brightness of a BGR colour."""
    b, g, r = bgr
    return 0.114 * b + 0.587 * g + 0.299 * r


def compute_team_remap(team_classifier, crops: list):
    """
    Map the classifier's arbitrary cluster ids (0/1) to *stable* team ids based
    on kit colour, so "Team 1" is the same physical team across runs/halves.

    Team 0 is fixed as the darker kit (lower luminance) -- a deterministic rule
    that orders the two clusters the same way every run as long as the kits are
    colour-separable. Returns (remap, colors) where remap[cluster] -> team_id
    and colors[team_id] is that team's median BGR kit colour.
    """
    labels = np.asarray(team_classifier.predict(crops))
    cluster_color = {}
    for c in (0, 1):
        idx = np.where(labels == c)[0]
        cluster_color[c] = (_dominant_color_bgr([crops[i] for i in idx])
                            if len(idx) else np.zeros(3))

    if _luminance_bgr(cluster_color[0]) <= _luminance_bgr(cluster_color[1]):
        remap = np.array([0, 1])   # cluster0 (darker) -> Team 0
    else:
        remap = np.array([1, 0])   # cluster1 (darker) -> Team 0

    colors = {int(remap[c]): cluster_color[c] for c in (0, 1)}
    return remap, colors


def final_third_zone(bx: float, by: float, attacks_right: bool,
                     config: SoccerPitchConfiguration):
    """
    Map a pitch-space ball position to a final-third key area, or None if the
    ball is not yet in the attacking final third.

    BOX     : inside the penalty box (near goal).
    CENTRAL : inner part of the box width, in the band beyond the box.
    HALF_L/R: outer strips of the box width, in that same band.
    WIDE_L/R: outside the box width anywhere in the final third.
    'L' = lower y, 'R' = higher y (absolute pitch sides).
    """
    goal_x = config.length if attacks_right else 0.0
    dist = abs(goal_x - bx)
    if dist > config.length * FINAL_THIRD_FRAC:
        return None
    # Penalty-box region (near goal)
    if dist <= BOX_DEPTH:
        if BOX_LO <= by <= BOX_HI:
            return "BOX"
        return "WIDE_L" if by < BOX_LO else "WIDE_R"
    # Outer band: box front edge -> final-third line
    if by < BOX_LO or by > BOX_HI:
        return "WIDE_L" if by < BOX_LO else "WIDE_R"
    if CEN_LO <= by <= CEN_HI:
        return "CENTRAL"
    return "HALF_L" if by < CEN_LO else "HALF_R"


def _zone_rects(goal_x: float, config: SoccerPitchConfiguration):
    """Return {zone: (x0, x1, y0, y1)} pitch-space rectangles for one end."""
    s = 1.0 if goal_x == config.length else -1.0
    near = goal_x - s * BOX_DEPTH                      # box front edge
    far = goal_x - s * (config.length * FINAL_THIRD_FRAC)  # final-third line
    W = config.width
    rects = {
        "BOX": (goal_x, near, BOX_LO, BOX_HI),
        "CENTRAL": (near, far, CEN_LO, CEN_HI),
        "HALF_L": (near, far, BOX_LO, CEN_LO),
        "HALF_R": (near, far, CEN_HI, BOX_HI),
        "WIDE_L": (goal_x, far, 0, BOX_LO),
        "WIDE_R": (goal_x, far, BOX_HI, W),
    }
    return rects, far


def draw_zone_overlay(pitch_img: np.ndarray, zone_entries: dict,
                      config: SoccerPitchConfiguration,
                      scale: float = 0.1, padding: int = 50,
                      team0_right: bool = TEAM0_ATTACKS_RIGHT) -> np.ndarray:
    """Shade the final-third key areas and draw each team's entry counts."""
    fills = {  # BGR
        "BOX": (230, 140, 40), "CENTRAL": (235, 180, 90),
        "HALF_L": (60, 170, 245), "HALF_R": (60, 170, 245),
        "WIDE_L": (90, 190, 110), "WIDE_R": (90, 190, 110),
    }

    def to_px(x, y):
        return int(x * scale) + padding, int(y * scale) + padding

    for goal_x in (0.0, config.length):
        attacks_right = (goal_x == config.length)
        team = 0 if team_attacks_right(0, team0_right) == attacks_right else 1
        rects, far = _zone_rects(goal_x, config)

        overlay = pitch_img.copy()
        for zname, (x0, x1, y0, y1) in rects.items():
            cv2.rectangle(overlay, to_px(min(x0, x1), min(y0, y1)),
                          to_px(max(x0, x1), max(y0, y1)), fills[zname], -1)
        cv2.addWeighted(overlay, 0.25, pitch_img, 0.75, 0, pitch_img)

        color = (255, 191, 0) if team == 0 else (147, 20, 255)
        for zname, (x0, x1, y0, y1) in rects.items():
            cx, cy = to_px((x0 + x1) / 2, (y0 + y1) / 2)
            cv2.putText(pitch_img, str(zone_entries[team][zname]),
                        (cx - 8, cy + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                        color, 2, cv2.LINE_AA)

        bx, _ = to_px(far, 0)
        cv2.line(pitch_img, (bx, padding),
                 (bx, int(config.width * scale) + padding),
                 (255, 255, 255), 1, cv2.LINE_AA)
    return pitch_img


def total_entries(zone_entries: dict, team: int) -> int:
    """Total ball entries across all zones for a team."""
    return int(sum(zone_entries[team].values()))


def format_clock(seconds: float) -> str:
    """Format a number of seconds as MM:SS."""
    return f"{int(seconds // 60):02d}:{int(seconds % 60):02d}"


def write_report(team_xg: dict, zone_entries: dict, source_desc: str,
                 clock_str: str, interval_history: list,
                 out_path: str) -> str:
    """
    Write the end-of-run match report to a text file, including Expected
    Goals (xG) and the final-third entries by key area. Returns the path.
    """
    lines = []
    lines.append("=" * 60)
    lines.append("REAL-TIME SOCCER ANALYTICS - MATCH REPORT")
    lines.append("=" * 60)
    lines.append(f"Generated : {time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Source    : {source_desc}")
    lines.append(f"Match time: {clock_str}")
    lines.append("")

    # Expected Goals
    lines.append("=== Expected Goals (xG) ===")
    lines.append(f"  Team 1: {team_xg[0]:.2f}")
    lines.append(f"  Team 2: {team_xg[1]:.2f}")
    lines.append("")

    # Final-third entries
    lines.append("=== Final-Third Entries by Key Area (possession-based) ===")
    lines.append("Columns: WIDE_L / HALF_L / CENTRAL / HALF_R / WIDE_R "
                 "(L = low y, R = high y)")
    for team in (0, 1):
        counts = "  ".join(f"{ZONE_LABELS[z]}:{zone_entries[team][z]}"
                           for z in ZONE_NAMES)
        lines.append(f"  Team {team + 1}:  {counts}")
        lines.append(f"           Total entries: "
                     f"{total_entries(zone_entries, team)}")
    lines.append("")

    # Interval history (xG + total entries per 15-min block)
    if interval_history:
        lines.append(f"=== Interval Reports (every {INTERVAL_MINUTES} min) ===")
        for rec in interval_history:
            t1, t2 = rec["teams"]["team1"], rec["teams"]["team2"]
            lines.append(
                f"  @ {rec['clock']}: "
                f"xG T1 {t1['xg']:.2f} / T2 {t2['xg']:.2f}  |  "
                f"entries T1 {t1['entries']} / T2 {t2['entries']}")
        lines.append("")

    report = "\n".join(lines) + "\n"
    with open(out_path, "w") as f:
        f.write(report)
    return out_path


def write_report_xlsx(team_xg: dict, zone_entries: dict, source_desc: str,
                      clock_str: str, interval_history: list,
                      all_shots: list, out_path: str) -> str:
    """
    Write the match report to an Excel workbook the coach can open directly.

    Sheets:
      - Summary  : match info, xG, and final-third entries per zone
      - Intervals: xG + entries per reporting block
      - Shots    : every logged shot candidate (accepted + rejected)

    Requires openpyxl (`pip install openpyxl`). Returns the path written.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    bold = Font(bold=True)
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Real-Time Soccer Analytics - Match Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Generated", time.strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Source", source_desc])
    ws.append(["Match time", clock_str])
    ws.append([])

    ws.append(["Expected Goals (xG)"])
    ws.cell(ws.max_row, 1).font = bold
    ws.append(["Team", "xG"])
    for c in ws[ws.max_row]:
        c.font = bold
    ws.append(["Team 1", round(team_xg[0], 2)])
    ws.append(["Team 2", round(team_xg[1], 2)])
    ws.append([])

    ws.append(["Final-Third Entries by Key Area (possession-based)"])
    ws.cell(ws.max_row, 1).font = bold
    header = ["Team"] + [ZONE_LABELS[z] for z in ZONE_NAMES] + ["Total"]
    ws.append(header)
    for c in ws[ws.max_row]:
        c.font = bold
    for team in (0, 1):
        row = ([f"Team {team + 1}"]
               + [zone_entries[team][z] for z in ZONE_NAMES]
               + [total_entries(zone_entries, team)])
        ws.append(row)

    # --- Intervals sheet ---
    wsi = wb.create_sheet("Intervals")
    wsi.append(["Clock", "xG Team 1", "xG Team 2",
                "Entries Team 1", "Entries Team 2"])
    for c in wsi[1]:
        c.font = bold
    for rec in interval_history:
        t1, t2 = rec["teams"]["team1"], rec["teams"]["team2"]
        wsi.append([rec["clock"], round(t1["xg"], 2), round(t2["xg"], 2),
                    t1["entries"], t2["entries"]])

    # --- Shots sheet ---
    wss = wb.create_sheet("Shots")
    fields = ["clock", "frame", "team", "accepted", "reason", "xg",
              "speed_mps", "dist_m", "aim_deg", "max_step_m", "x_m", "y_m"]
    wss.append(fields)
    for c in wss[1]:
        c.font = bold
    for s in all_shots:
        wss.append([s.get(k) for k in fields])

    # Auto-size columns for readability.
    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None),
                        default=8)
            sheet.column_dimensions[col[0].column_letter].width = width + 2

    wb.save(out_path)
    return out_path


def print_interval_report(interval_idx: int, clock_str: str,
                          team_xg: dict, zone_entries: dict,
                          prev: dict) -> dict:
    """Print a per-interval coach update and return it as a record."""
    print(f"\n===== Coach update @ {clock_str} "
          f"(interval {interval_idx}, last {INTERVAL_MINUTES} min) =====")
    record = {"interval": interval_idx, "clock": clock_str, "teams": {}}
    for team in (0, 1):
        tot = total_entries(zone_entries, team)
        d_xg = team_xg[team] - prev['xg'][team]
        d_ent = tot - prev['entries'][team]
        print(f"  Team {team + 1}: xG {team_xg[team]:.2f} (+{d_xg:.2f})  |  "
              f"entries {tot} (+{d_ent})")
        record["teams"][f"team{team + 1}"] = {
            "xg": round(team_xg[team], 2), "xg_delta": round(d_xg, 2),
            "entries": tot, "entries_delta": d_ent,
        }
        prev['xg'][team] = team_xg[team]
        prev['entries'][team] = tot
    return record


def build_broadcast_payload(team_xg: dict, zone_entries: dict, clock: str,
                            interval_history: list) -> dict:
    """Assemble the JSON payload broadcast to the coach dashboard."""
    teams = {}
    for team in (0, 1):
        teams[f"team{team + 1}"] = {
            "xg": round(team_xg[team], 2),
            "entries": {z: zone_entries[team][z] for z in ZONE_NAMES},
            "total_entries": total_entries(zone_entries, team),
        }
    return {
        "clock": clock,
        "updated": time.time(),
        "zone_order": ZONE_NAMES,
        "zone_labels": ZONE_LABELS,
        "teams": teams,
        "intervals": interval_history,
    }


def draw_dashboard(team_xg: dict, zone_entries: dict,
                   match_time: str = "") -> np.ndarray:
    """
    Render a coach-facing dashboard with the two new metrics:
    cumulative Expected Goals (xG) and ball entries by zone.
    """
    W, H = 560, 640
    dash = np.full((H, W, 3), 28, dtype=np.uint8)

    cyan = (255, 191, 0)     # Team 1 (BGR)
    pink = (147, 20, 255)    # Team 2 (BGR)
    white = (255, 255, 255)
    gray = (165, 165, 165)
    font = cv2.FONT_HERSHEY_SIMPLEX

    def text(s, org, scale, color, thick=1):
        cv2.putText(dash, s, org, font, scale, color, thick, cv2.LINE_AA)

    # Title
    text("COACH DASHBOARD", (20, 42), 0.9, white, 2)
    if match_time:
        text(match_time, (W - 120, 42), 0.8, (0, 255, 255), 2)
    cv2.line(dash, (20, 58), (W - 20, 58), gray, 1)

    # Team column headers
    text("TEAM 1", (165, 96), 0.75, cyan, 2)
    text("TEAM 2", (390, 96), 0.75, pink, 2)

    # --- Expected Goals ---
    text("EXPECTED GOALS (xG)", (20, 140), 0.6, gray, 1)
    text(f"{team_xg[0]:.2f}", (165, 195), 1.3, cyan, 3)
    text(f"{team_xg[1]:.2f}", (390, 195), 1.3, pink, 3)
    cv2.line(dash, (20, 225), (W - 20, 225), (60, 60, 60), 1)

    # --- Final-third entries by key area ---
    text("FINAL-THIRD ENTRIES", (20, 262), 0.6, gray, 1)
    text("L = low y  -  R = high y", (20, 286), 0.4, gray, 1)

    row_y0, row_dy = 320, 40
    totals = {0: 0, 1: 0}
    for i, zname in enumerate(ZONE_NAMES):
        y = row_y0 + i * row_dy
        text(ZONE_LABELS[zname], (20, y), 0.55, white, 1)
        for team, color, x in ((0, cyan, 170), (1, pink, 395)):
            c = zone_entries[team][zname]
            totals[team] += c
            text(str(c), (x, y), 0.6, color, 2)

    # Totals
    ty = row_y0 + len(ZONE_NAMES) * row_dy + 8
    cv2.line(dash, (20, ty - 28), (W - 20, ty - 28), (60, 60, 60), 1)
    text("TOTAL", (20, ty), 0.6, white, 2)
    text(str(totals[0]), (170, ty), 0.6, cyan, 2)
    text(str(totals[1]), (395, ty), 0.6, pink, 2)

    return dash


def draw_pitch_voronoi_diagram_2(
    config: SoccerPitchConfiguration,
    team_1_xy: np.ndarray,
    team_2_xy: np.ndarray,
    team_1_color: sv.Color = sv.Color.RED,
    team_2_color: sv.Color = sv.Color.WHITE,
    opacity: float = 0.5,
    padding: int = 50,
    scale: float = 0.1,
    pitch: Optional[np.ndarray] = None
) -> np.ndarray:
    """
    Draws a Voronoi diagram on a soccer pitch representing the control areas of two
    teams with smooth color transitions.
    """
    if pitch is None:
        pitch = draw_pitch(
            config=config,
            padding=padding,
            scale=scale
        )

    scaled_width = int(config.width * scale)
    scaled_length = int(config.length * scale)

    voronoi = np.zeros_like(pitch, dtype=np.uint8)

    team_1_color_bgr = np.array(team_1_color.as_bgr(), dtype=np.uint8)
    team_2_color_bgr = np.array(team_2_color.as_bgr(), dtype=np.uint8)

    y_coordinates, x_coordinates = np.indices((
        scaled_width + 2 * padding,
        scaled_length + 2 * padding
    ))

    y_coordinates -= padding
    x_coordinates -= padding

    def calculate_distances(xy, x_coords, y_coords):
        return np.sqrt(
            (xy[:, 0][:, None, None] * scale - x_coords) ** 2 +
            (xy[:, 1][:, None, None] * scale - y_coords) ** 2
        )

    distances_team_1 = calculate_distances(team_1_xy, x_coordinates, y_coordinates)
    distances_team_2 = calculate_distances(team_2_xy, x_coordinates, y_coordinates)

    min_distances_team_1 = np.min(distances_team_1, axis=0)
    min_distances_team_2 = np.min(distances_team_2, axis=0)

    steepness = 15
    distance_ratio = min_distances_team_2 / np.clip(
        min_distances_team_1 + min_distances_team_2, a_min=1e-5, a_max=None)
    blend_factor = np.tanh((distance_ratio - 0.5) * steepness) * 0.5 + 0.5

    for c in range(3):
        voronoi[:, :, c] = (
            blend_factor * team_1_color_bgr[c] +
            (1 - blend_factor) * team_2_color_bgr[c]
        ).astype(np.uint8)

    overlay = cv2.addWeighted(voronoi, opacity, pitch, 1 - opacity, 0)
    return overlay



# CAMERA SELECTION

def resolve_video_source(source: str) -> str:
    """
    Resolve a video source to something OpenCV can open.

    - Local file paths are returned unchanged.
    - Direct stream URLs (HLS .m3u8, DASH .mpd, RTSP, RTMP) are returned
      unchanged so OpenCV's FFmpeg backend opens them straight away. This is
      the path used for Veo Live, which serves the game as an HLS stream.
    - Other http(s) URLs (including YouTube) are resolved to a direct stream
      URL using yt-dlp, preferring a progressive MP4 ≤720p for stable decoding.
    """
    low = source.lower()

    # Non-http sources (local files, rtsp://, rtmp://) pass through unchanged.
    if not low.startswith(("http://", "https://")):
        return source

    # Direct streaming manifests (e.g. Veo Live HLS) are opened directly by
    # OpenCV/FFmpeg; sending these to yt-dlp would fail.
    direct_markers = (".m3u8", ".mpd")
    if any(m in low for m in direct_markers):
        print(f"📡 Using direct live stream (HLS/DASH): {source}")
        return source

    try:
        import yt_dlp
    except ImportError:
        raise RuntimeError(
            "yt-dlp is required for URL/YouTube input. "
            "Install it with: pip install yt-dlp")

    print(f"🔗 Resolving stream URL via yt-dlp: {source}")
    ydl_opts = {
        "quiet": True,
        "no_warnings": True,
        # Prefer a single progressive MP4 stream <=720p (no separate audio)
        "format": "best[ext=mp4][height<=720]/best[height<=720]/best",
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info = ydl.extract_info(source, download=False)
        # For playlists, take the first entry
        if "entries" in info:
            info = info["entries"][0]
        stream_url = info.get("url")

    if not stream_url:
        raise RuntimeError(f"Could not resolve a stream URL for: {source}")

    print("✅ Stream URL resolved")
    return stream_url


def select_camera() -> int:
    """Scan available cameras and let the user pick one."""
    available = []
    for i in range(10):
        test_cap = cv2.VideoCapture(i)
        if test_cap.isOpened():
            ret, frame = test_cap.read()
            if ret:
                h, w = frame.shape[:2]
                available.append((i, w, h))
            test_cap.release()

    if not available:
        raise RuntimeError("No cameras found")

    print("\n📷 Available cameras:")
    for idx, (cam_id, w, h) in enumerate(available):
        print(f"  [{idx}] Camera {cam_id}  ({w}x{h})")

    choice = input("\nSelect camera number [0]: ").strip()
    choice = int(choice) if choice else 0
    return available[choice][0]



# CALIBRATION (collect crops, fit TeamClassifier)

def get_device() -> str:
    """Pick the best available torch device: Apple MPS, CUDA, else CPU."""
    mps = getattr(torch.backends, "mps", None)
    if mps is not None and mps.is_available():
        return "mps"
    if torch.cuda.is_available():
        return "cuda"
    return "cpu"


MIN_CROPS = 20  # UMAP needs at least n_neighbors+1 samples (default 15)

def _collect_crops_from_frame(frame: np.ndarray) -> list:
    """Detect players in a frame and return their torso/jersey crops."""
    result = PLAYER_DETECTION_MODEL.infer(frame, confidence=0.3)[0]
    detections = sv.Detections.from_inference(result)
    detections = detections.with_nms(threshold=0.5, class_agnostic=True)
    players = detections[detections.class_id == PLAYER_ID]
    return [jersey_crop(frame, xyxy) for xyxy in players.xyxy]


def calibrate_team_classifier(cap: cv2.VideoCapture,
                              seekable: bool = False) -> TeamClassifier:
    """
    Collect player crops and fit the SiGLIP-based TeamClassifier.

    When the source is seekable (a video file), sample frames spread evenly
    across the whole match for a representative fit. Otherwise (live streams),
    read a sequential burst from the current position.
    """
    crops = []
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) if seekable else 0

    if seekable and total > CALIBRATION_SAMPLES:
        print(f"\n🔄 Calibrating team classifier "
              f"({CALIBRATION_SAMPLES} frames spread across the match)...")
        indices = np.linspace(0, total - 1, CALIBRATION_SAMPLES).astype(int)
        for n, idx in enumerate(indices, start=1):
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(idx))
            ret, frame = cap.read()
            if not ret:
                continue
            crops += _collect_crops_from_frame(frame)
            if n % 20 == 0:
                print(f"   … sampled {n}/{CALIBRATION_SAMPLES} frames, "
                      f"{len(crops)} crops so far")
    else:
        print(f"\n🔄 Calibrating team classifier "
              f"(up to {CALIBRATION_FRAMES} frames, need ≥{MIN_CROPS} crops)...")
        frame_count = 0
        while frame_count < CALIBRATION_FRAMES:
            ret, frame = cap.read()
            if not ret:
                break
            frame_count += 1
            if frame_count % CALIBRATION_STRIDE != 0:
                continue
            crops += _collect_crops_from_frame(frame)
            if frame_count % 30 == 0:
                print(f"   … frame {frame_count}, {len(crops)} crops so far")

    # UMAP needs enough samples; pad by duplicating existing crops if short
    if len(crops) < MIN_CROPS:
        if len(crops) == 0:
            print("⚠️  No player crops found. Using blank fallback crops.")
            crops = [np.zeros((64, 32, 3), dtype=np.uint8)] * MIN_CROPS
        else:
            print(f"⚠️  Only {len(crops)} crops — duplicating to reach {MIN_CROPS}")
            while len(crops) < MIN_CROPS:
                crops.append(crops[len(crops) % len(crops)])

    device = get_device()
    print(f"🖥️  Team classifier device: {device}")
    team_classifier = TeamClassifier(device=device)
    team_classifier.fit(crops)

    # Map arbitrary cluster ids to stable, colour-based team ids.
    team_remap, colors = compute_team_remap(team_classifier, crops)
    print(f"✅ Team classifier fitted on {len(crops)} crops")
    for t in (0, 1):
        b, g, r = (int(v) for v in colors.get(t, np.zeros(3)))
        print(f"🎨 Team {t + 1} kit colour ~ RGB({r},{g},{b})")
    print("   (Team 1 = darker kit. Use --swap-teams if this is reversed.)")
    return team_classifier, team_remap



# LIVE FRAME GRABBER

class FrameGrabber:
    """
    Background reader that always keeps only the most recent frame.

    For live sources (RTSP/RTMP/HLS), decoding can lag behind real time. A
    plain cap.read() then drains a growing buffer, so latency keeps climbing.
    This thread continuously pulls frames and discards stale ones, so the
    main loop always processes the freshest frame (constant, low latency at
    the cost of dropping frames the analytics can't keep up with).
    """

    def __init__(self, cap: cv2.VideoCapture):
        self.cap = cap
        self.lock = threading.Lock()
        self.frame = None
        self.ret = False
        self.ended = False
        self.stopped = False
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def _run(self):
        while not self.stopped:
            ret, frame = self.cap.read()
            if not ret:
                self.ended = True
                break
            with self.lock:
                self.ret, self.frame = True, frame

    def read(self):
        """Return (ret, frame) for the latest frame, waiting briefly if none yet."""
        for _ in range(int(LIVE_MAX_MISSES)):
            with self.lock:
                if self.frame is not None:
                    frame, self.frame = self.frame, None  # consume once
                    return True, frame
            if self.ended:
                return False, None
            time.sleep(LIVE_RETRY_DELAY)
        return (False, None) if self.ended else (False, None)

    def release(self):
        self.stopped = True
        self.thread.join(timeout=1.0)
        self.cap.release()


# MAIN

def main(video_path: Optional[str] = None, realtime: bool = False,
         stride: int = 1, report_path: Optional[str] = None,
         no_display: bool = False, flip_attack: bool = False,
         swap_teams: bool = False):

    show = not no_display  # whether to render GUI windows / visualizations
    # Whether Team 0 attacks the x=length goal this run (flips at half-time).
    team0_right = TEAM0_ATTACKS_RIGHT != flip_attack

    # Source: video file / URL (YouTube) or live camera
    is_url = bool(video_path) and video_path.lower().startswith(
        ("http://", "https://", "rtsp://", "rtmp://"))
    # Live streams (Veo Live HLS/DASH, RTSP/RTMP) are not seekable and can
    # briefly return no frame without having ended.
    is_live = bool(video_path) and (
        any(m in video_path.lower() for m in (".m3u8", ".mpd"))
        or video_path.lower().startswith(("rtsp://", "rtmp://")))
    if video_path:
        if not is_url and not os.path.exists(video_path):
            print(f"❌ Video file not found: {video_path}")
            return
        try:
            resolved = resolve_video_source(video_path)
        except RuntimeError as e:
            print(f"❌ {e}")
            return
        print(f"🎞️  Using video source: {video_path}")
        # Force the FFmpeg backend for network sources for reliable HLS/RTSP.
        cap = (cv2.VideoCapture(resolved, cv2.CAP_FFMPEG)
               if is_url else cv2.VideoCapture(resolved))
        label = video_path if is_url else os.path.basename(video_path)
        source_desc = f"live stream '{label}'" if is_live else f"video '{label}'"
    else:
        cam_id = select_camera()
        cap = cv2.VideoCapture(cam_id)
        source_desc = f"Camera {cam_id}"

    if not cap.isOpened():
        print("❌ Could not open video source")
        return

    # Calibrate team classifier. Local files are seekable, so sample frames
    # spread across the whole match for a representative team-colour fit.
    seekable = bool(video_path) and not is_url
    team_classifier, team_remap = calibrate_team_classifier(
        cap, seekable=seekable)
    if swap_teams:
        team_remap = 1 - team_remap
        print("🔁 --swap-teams: Team 1 and Team 2 identities swapped.")
    print(f"🧭 Team 1 attacks {'right (x=length)' if team0_right else 'left (x=0)'} "
          f"this run.")

    # Rewind local video so calibration frames are not skipped.
    # (Network/YouTube streams are usually not seekable, so skip the rewind.)
    if seekable:
        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

    # Annotators
    ellipse_annotator = sv.EllipseAnnotator(
        color=sv.ColorPalette.from_hex(['#00BFFF', '#FF1493', '#FFD700']),
        thickness=2
    )
    label_annotator = sv.LabelAnnotator(
        color=sv.ColorPalette.from_hex(['#00BFFF', '#FF1493', '#FFD700']),
        text_color=sv.Color.from_hex('#000000'),
        text_position=sv.Position.BOTTOM_CENTER
    )
    triangle_annotator = sv.TriangleAnnotator(
        color=sv.Color.from_hex('#FFD700'),
        base=20,
        height=17
    )

    # Tracker
    tracker = sv.ByteTrack()
    tracker.reset()

    # Homography smoothing
    M_buffer = deque(maxlen=MAXLEN)
    # Last good view transform; refreshed periodically and reused in between
    # (see FIELD_DETECT_INTERVAL) instead of re-detecting keypoints every frame.
    transformer = None

    # Team-id cache keyed by tracker id. `team_votes` accumulates per-track
    # SiGLIP votes; once a track reaches TEAM_VOTE_FRAMES it is locked into
    # `team_by_id` and the classifier is never run on it again.
    team_by_id: dict = {}
    team_votes = defaultdict(lambda: [0, 0])

    # Rolling buffer of recent jersey crops for on-demand re-calibration ('c')
    recent_crops = deque(maxlen=RECAL_BUFFER)

    # Expected Goals (xG) state
    ball_history = deque(maxlen=BALL_HISTORY_LEN)
    team_xg = {0: 0.0, 1: 0.0}
    shot_markers = []          # list of dicts: {frame, xg, position, team}
    all_shots = []             # unpruned record of every detected shot (for CSV)
    last_shot_frame = -SHOT_COOLDOWN_FRAMES
    frame_count = 0
    processed_idx = 0   # number of frames actually processed (drives throttling)

    # Zone-entry state: counts[team][(half, lane)] and each team's last zone
    zone_entries = {0: defaultdict(int), 1: defaultdict(int)}
    # Debounced zone-entry state: an entry is confirmed only after the same
    # team holds the ball in the same zone for ZONE_CONFIRM_FRAMES frames.
    confirmed_zone = {0: None, 1: None}   # last confirmed zone per team
    pending_zone = {0: None, 1: None}     # candidate zone awaiting confirmation
    pending_count = {0: 0, 1: 0}          # consecutive frames in the candidate
    last_poss_team = None

    # Match-clock / interval-reporting state
    fps = cap.get(cv2.CAP_PROP_FPS)
    fps = fps if fps and fps > 1 else 30.0
    interval_frames = max(int(INTERVAL_MINUTES * 60 * fps), 1)
    next_interval = interval_frames
    interval_idx = 0
    prev_snapshot = {'xg': {0: 0.0, 1: 0.0}, 'entries': {0: 0, 1: 0}}
    interval_history: list = []

    # Live broadcast to the coach dashboard (no-op unless FIREBASE_DB_URL set)
    broadcaster = StatsBroadcaster.from_env()

    print(f"\n{source_desc} started. Press Q to quit, C to re-calibrate teams")
    print(f"⏱️  Video ~{fps:.1f} fps; coach updates every "
          f"{INTERVAL_MINUTES} min of match time\n")

    live_misses = 0
    while True:
        # Advance `stride` frames (cheaply skipping the in-between ones with
        # grab()) and process only the most recent one. stride>1 trades
        # temporal resolution for a near-linear speedup -- ideal for
        # batch-processing a full half. frame_count tracks true video frames so
        # the match clock and time-based logic stay accurate.
        grabbed = False
        for _ in range(max(1, stride)):
            if not cap.grab():
                break
            frame_count += 1
            grabbed = True

        if not grabbed:
            # Live streams can stall briefly (buffer underrun); retry a few
            # times before giving up. Files/VOD end immediately on EOF.
            if is_live and live_misses < LIVE_MAX_MISSES:
                live_misses += 1
                time.sleep(LIVE_RETRY_DELAY)
                continue
            break
        live_misses = 0

        ret, frame = cap.retrieve()
        if not ret:
            break
        processed_idx += 1

        # DETECTION
        result = PLAYER_DETECTION_MODEL.infer(frame, confidence=0.3)[0]
        detections = sv.Detections.from_inference(result)

        # Ball
        ball_detections = detections[detections.class_id == BALL_ID]
        if len(ball_detections) > MAX_BALLS:
            ball_detections = ball_detections[:MAX_BALLS]
        ball_detections.xyxy = sv.pad_boxes(ball_detections.xyxy, px=10)

        # Others (goalkeeper, player, referee). Track FIRST on stable role
        # class ids, then assign teams -- this lets us cache each player's team
        # by tracker id instead of re-running the SiGLIP classifier on every
        # crop every frame.
        all_detections = detections[detections.class_id != BALL_ID]
        all_detections = all_detections.with_nms(threshold=0.5, class_agnostic=True)
        all_detections.class_id = all_detections.class_id.astype(int)

        tracked = tracker.update_with_detections(detections=all_detections)

        goalkeepers_detections = tracked[
            tracked.class_id == GOALKEEPER_ID][:MAX_GOALKEEPERS]
        players_detections = tracked[
            tracked.class_id == PLAYER_ID][:MAX_PLAYERS]
        referees_detections = tracked[
            tracked.class_id == REFEREE_ID][:MAX_REFEREES]


        # TEAM ASSIGNMENT (cached by tracker id)
        # Only embed crops for tracks that have not yet locked a team; once a
        # track reaches TEAM_VOTE_FRAMES votes its team is fixed, so the
        # expensive SiGLIP call is skipped for it on every subsequent frame.
        if len(players_detections) > 0:
            player_tids = players_detections.tracker_id
            crops_to_class, meta = [], []
            for i, tid in enumerate(player_tids):
                if int(tid) in team_by_id:
                    continue
                crops_to_class.append(
                    jersey_crop(frame, players_detections.xyxy[i]))
                meta.append((i, int(tid)))

            if crops_to_class:
                preds = team_remap[np.asarray(
                    team_classifier.predict(crops_to_class))]
                recent_crops.extend(crops_to_class)
                for (i, tid), pred in zip(meta, preds):
                    votes = team_votes[tid]
                    votes[int(pred)] += 1
                    if sum(votes) >= TEAM_VOTE_FRAMES:
                        team_by_id[tid] = 0 if votes[0] >= votes[1] else 1

            team_ids = np.empty(len(player_tids), dtype=int)
            for i, tid in enumerate(player_tids):
                tid = int(tid)
                if tid in team_by_id:
                    team_ids[i] = team_by_id[tid]
                else:
                    v = team_votes[tid]
                    team_ids[i] = 0 if v[0] >= v[1] else 1   # provisional vote
            players_detections.class_id = team_ids

        goalkeepers_detections.class_id = resolve_goalkeepers_team_id(
            players_detections, goalkeepers_detections)

        # Referees -> palette index 2 (gold).
        referees_detections.class_id = np.full(
            len(referees_detections), 2, dtype=int)

        # MERGE for annotation (tracker ids preserved for labels).
        merged_detections = sv.Detections.merge([
            players_detections, goalkeepers_detections, referees_detections
        ])
        merged_detections.class_id = merged_detections.class_id.astype(int)

        labels = [
            f"#{tracker_id}"
            for tracker_id in merged_detections.tracker_id
        ]


        # ANNOTATE CAMERA FRAME (only needed when displaying windows)
        if show:
            annotated_frame = frame.copy()
            annotated_frame = ellipse_annotator.annotate(
                scene=annotated_frame, detections=merged_detections)
            annotated_frame = label_annotator.annotate(
                scene=annotated_frame, detections=merged_detections, labels=labels)
            annotated_frame = triangle_annotator.annotate(
                scene=annotated_frame, detections=ball_detections)


        # FIELD KEYPOINTS + HOMOGRAPHY
        # Re-detecting keypoints every frame is expensive and largely redundant
        # (slow camera pan + already-smoothed homography). Refresh only every
        # FIELD_DETECT_INTERVAL processed frames and reuse the last good
        # transform in between.
        if transformer is None or processed_idx % FIELD_DETECT_INTERVAL == 0:
            result_field = FIELD_DETECTION_MODEL.infer(frame, confidence=0.3)[0]
            keypoints = sv.KeyPoints.from_inference(result_field)

            have_keypoints = not (
                keypoints.xy is None or len(keypoints.xy) == 0
                or keypoints.confidence is None)
            if have_keypoints:
                filter_mask = keypoints.confidence[0] > 0.5
                frame_reference_points = keypoints.xy[0][filter_mask]
                pitch_reference_points = np.array(CONFIG.vertices)[filter_mask]
                if len(frame_reference_points) >= 4:
                    new_transformer = ViewTransformer(
                        source=frame_reference_points,
                        target=pitch_reference_points
                    )
                    # Smooth homography across recent detections.
                    M_buffer.append(new_transformer.m)
                    new_transformer.m = np.mean(np.array(M_buffer), axis=0)
                    transformer = new_transformer

        # Until the first successful homography fix we cannot project to pitch.
        if transformer is None:
            if show:
                cv2.imshow("Camera View", annotated_frame)
                if cv2.waitKey(1) & 0xFF == ord("q"):
                    break
            continue


        # PROJECT TO PITCH
        # Merge players + goalkeepers for pitch projection
        pitch_detections = sv.Detections.merge([
            players_detections, goalkeepers_detections
        ])

        if len(pitch_detections) == 0 and len(ball_detections) == 0:
            if show:
                cv2.imshow("Camera View", annotated_frame)
                if cv2.waitKey(1) & 0xFF == ord("q"):
                    break
            continue

        # Ball
        frame_ball_xy = ball_detections.get_anchors_coordinates(
            sv.Position.BOTTOM_CENTER)
        pitch_ball_xy = transformer.transform_points(
            points=frame_ball_xy) if len(frame_ball_xy) > 0 else np.array([])

        # Players + goalkeepers
        players_xy = pitch_detections.get_anchors_coordinates(
            sv.Position.BOTTOM_CENTER)
        pitch_players_xy = transformer.transform_points(
            points=players_xy) if len(players_xy) > 0 else np.array([])

        # Referees
        referees_xy = referees_detections.get_anchors_coordinates(
            sv.Position.BOTTOM_CENTER)
        pitch_referees_xy = transformer.transform_points(
            points=referees_xy) if len(referees_xy) > 0 else np.array([])

        # EXPECTED GOALS (xG)
        if len(pitch_ball_xy) > 0:
            ball_history.append(pitch_ball_xy[0])

            is_shot, vx, info = detect_shot(ball_history, CONFIG,
                                            frame_gap=stride, fps=fps)
            cooldown_ok = (frame_count - last_shot_frame) >= SHOT_COOLDOWN_FRAMES
            accepted = is_shot and cooldown_ok

            # Log every loose candidate (accepted or not) for offline tuning.
            if info.get("candidate") and len(pitch_detections) > 0:
                goal_x = CONFIG.length if vx >= 0 else 0.0
                bx, by = pitch_ball_xy[0]
                xg = calculate_xg(bx, by, goal_x, CONFIG)
                team = nearest_team(pitch_ball_xy[0], pitch_detections)
                all_shots.append({
                    'clock': format_clock(frame_count / fps),
                    'frame': frame_count,
                    'team': team + 1,
                    'accepted': int(accepted),
                    'reason': info["reason"],
                    'xg': round(xg, 3),
                    'speed_mps': info["speed_mps"],
                    'dist_m': info["dist_m"],
                    'aim_deg': info["aim_deg"],
                    'max_step_m': info["max_step_m"],
                    'x_m': round(bx / 100.0, 2),
                    'y_m': round(by / 100.0, 2),
                })
                if accepted:
                    team_xg[team] += xg
                    shot_markers.append({
                        'frame': frame_count,
                        'xg': xg,
                        'position': pitch_ball_xy[0].copy(),
                        'team': team,
                    })
                    last_shot_frame = frame_count
                    print(f"SHOT! Team {team + 1} xG: {xg:.2f} "
                          f"(total {team_xg[team]:.2f})")

        # ZONE ENTRIES (final-third key areas, possession-based, debounced)
        if len(pitch_ball_xy) > 0 and len(pitch_detections) > 0:
            poss_team = nearest_team(pitch_ball_xy[0], pitch_detections)
            zone = final_third_zone(pitch_ball_xy[0][0], pitch_ball_xy[0][1],
                                    team_attacks_right(poss_team, team0_right),
                                    CONFIG)

            # Restart the dwell whenever possession switches or the candidate
            # zone changes; otherwise extend the current dwell.
            if poss_team != last_poss_team or zone != pending_zone[poss_team]:
                pending_zone[poss_team] = zone
                pending_count[poss_team] = 1
            else:
                pending_count[poss_team] += 1
            last_poss_team = poss_team

            # Confirm the entry only after a sustained dwell in a real zone.
            if (pending_count[poss_team] >= ZONE_CONFIRM_FRAMES
                    and confirmed_zone[poss_team] != zone):
                confirmed_zone[poss_team] = zone
                if zone is not None:
                    zone_entries[poss_team][zone] += 1
                    print(f"Team {poss_team + 1} final-third entry -> "
                          f"{ZONE_LABELS[zone]}")

        # COACH INTERVAL REPORT (every INTERVAL_MINUTES of match time)
        if frame_count >= next_interval:
            interval_idx += 1
            record = print_interval_report(
                interval_idx, format_clock(frame_count / fps),
                team_xg, zone_entries, prev_snapshot)
            interval_history.append(record)
            next_interval += interval_frames

        # BROADCAST live stats to the coach dashboard (throttled internally)
        broadcaster.update(build_broadcast_payload(
            team_xg, zone_entries, format_clock(frame_count / fps),
            interval_history))


        # In headless mode skip all rendering/visualization (the expensive
        # radar + Voronoi drawing and the GUI windows) -- metrics above are
        # already computed and broadcast.
        if not show:
            continue

        # DRAW RADAR PITCH
        pitch_img = draw_pitch(CONFIG)

        # Ball
        if len(pitch_ball_xy) > 0:
            pitch_img = draw_points_on_pitch(
                config=CONFIG,
                xy=pitch_ball_xy,
                face_color=sv.Color.WHITE,
                edge_color=sv.Color.BLACK,
                radius=10,
                pitch=pitch_img)

        # Team 1
        if len(pitch_players_xy) > 0 and len(pitch_detections) > 0:
            team0_mask = pitch_detections.class_id == 0
            if team0_mask.any():
                pitch_img = draw_points_on_pitch(
                    config=CONFIG,
                    xy=pitch_players_xy[team0_mask],
                    face_color=sv.Color.from_hex('00BFFF'),
                    edge_color=sv.Color.BLACK,
                    radius=16,
                    pitch=pitch_img)

        # Team 2
        if len(pitch_players_xy) > 0 and len(pitch_detections) > 0:
            team1_mask = pitch_detections.class_id == 1
            if team1_mask.any():
                pitch_img = draw_points_on_pitch(
                    config=CONFIG,
                    xy=pitch_players_xy[team1_mask],
                    face_color=sv.Color.from_hex('FF1493'),
                    edge_color=sv.Color.BLACK,
                    radius=16,
                    pitch=pitch_img)

        # Referees
        if len(pitch_referees_xy) > 0:
            pitch_img = draw_points_on_pitch(
                config=CONFIG,
                xy=pitch_referees_xy,
                face_color=sv.Color.from_hex('FFD700'),
                edge_color=sv.Color.BLACK,
                radius=16,
                pitch=pitch_img)

        # Zone grid + per-zone entry counts
        pitch_img = draw_zone_overlay(pitch_img, zone_entries, CONFIG,
                                      team0_right=team0_right)

        # Recent shot markers on the radar (pitch is drawn at scale=0.1,
        # padding=50 by draw_pitch defaults)
        for shot in shot_markers:
            if frame_count - shot['frame'] < XG_DISPLAY_DURATION:
                px = int(shot['position'][0] * 0.1) + 50
                py = int(shot['position'][1] * 0.1) + 50
                cv2.putText(pitch_img, f"xG {shot['xg']:.2f}",
                            (px + 8, py), cv2.FONT_HERSHEY_SIMPLEX,
                            0.6, (255, 255, 255), 2, cv2.LINE_AA)
                cv2.circle(pitch_img, (px, py), 8, (0, 0, 255), 2)

        # Prune old markers
        shot_markers = [
            s for s in shot_markers
            if frame_count - s['frame'] < XG_DISPLAY_DURATION
        ]


        # VORONOI (smooth blend)
        if (len(pitch_players_xy) > 0 and len(pitch_detections) > 0):
            team0_mask = pitch_detections.class_id == 0
            team1_mask = pitch_detections.class_id == 1
            if team0_mask.any() and team1_mask.any():
                voronoi_img = draw_pitch(
                    config=CONFIG,
                    background_color=sv.Color.WHITE,
                    line_color=sv.Color.BLACK
                )
                voronoi_img = draw_pitch_voronoi_diagram_2(
                    config=CONFIG,
                    team_1_xy=pitch_players_xy[team0_mask],
                    team_2_xy=pitch_players_xy[team1_mask],
                    team_1_color=sv.Color.from_hex('00BFFF'),
                    team_2_color=sv.Color.from_hex('FF1493'),
                    pitch=voronoi_img)
                # Ball on voronoi
                if len(pitch_ball_xy) > 0:
                    voronoi_img = draw_points_on_pitch(
                        config=CONFIG,
                        xy=pitch_ball_xy,
                        face_color=sv.Color.WHITE,
                        edge_color=sv.Color.WHITE,
                        radius=8,
                        thickness=1,
                        pitch=voronoi_img)
                # Team dots on voronoi
                voronoi_img = draw_points_on_pitch(
                    config=CONFIG,
                    xy=pitch_players_xy[team0_mask],
                    face_color=sv.Color.from_hex('00BFFF'),
                    edge_color=sv.Color.WHITE,
                    radius=16,
                    thickness=1,
                    pitch=voronoi_img)
                voronoi_img = draw_points_on_pitch(
                    config=CONFIG,
                    xy=pitch_players_xy[team1_mask],
                    face_color=sv.Color.from_hex('FF1493'),
                    edge_color=sv.Color.WHITE,
                    radius=16,
                    thickness=1,
                    pitch=voronoi_img)

                cv2.imshow("Voronoi", voronoi_img)


        # CUMULATIVE xG OVERLAY (camera frame)
        cv2.putText(annotated_frame, f"Team 1 xG: {team_xg[0]:.2f}",
                    (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                    (255, 191, 0), 2, cv2.LINE_AA)
        cv2.putText(annotated_frame, f"Team 2 xG: {team_xg[1]:.2f}",
                    (10, 62), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                    (147, 20, 255), 2, cv2.LINE_AA)

        # COACH DASHBOARD (xG + zone entries)
        dashboard = draw_dashboard(team_xg, zone_entries,
                                   format_clock(frame_count / fps))

        # SHOW WINDOWS
        cv2.imshow("Camera View", annotated_frame)
        cv2.imshow("Pitch Radar", pitch_img)
        cv2.imshow("Coach Dashboard", dashboard)

        key = cv2.waitKey(1) & 0xFF
        if key == ord("q"):
            break
        if key == ord("c"):
            # Re-fit team colours on recently seen players (handles drift /
            # lighting changes, or a bad initial calibration).
            if len(recent_crops) >= MIN_CROPS:
                print(f"\n🔄 Re-calibrating on {len(recent_crops)} recent "
                      f"crops...")
                try:
                    crops_now = list(recent_crops)
                    team_classifier.fit(crops_now)
                    # Recompute the stable colour-based mapping (+ swap override).
                    team_remap, _ = compute_team_remap(team_classifier, crops_now)
                    if swap_teams:
                        team_remap = 1 - team_remap
                    print("✅ Team classifier re-fitted")
                except Exception as e:  # noqa: BLE001 - keep the match running
                    print(f"⚠️  Re-calibration failed ({e}); keeping previous fit")
            else:
                print(f"⚠️  Need ≥{MIN_CROPS} recent crops to re-calibrate "
                      f"(have {len(recent_crops)})")

    cap.release()
    cv2.destroyAllWindows()
    broadcaster.stop()

    # Expected Goals summary
    print("\n=== Expected Goals (xG) ===")
    print(f"  Team 1: {team_xg[0]:.2f}    Team 2: {team_xg[1]:.2f}")

    # Zone-entry summary (final-third key areas, possession-based)
    print("\n=== Final-Third Entries by Key Area (possession-based) ===")
    print("Columns: WIDE_L / HALF_L / CENTRAL / HALF_R / WIDE_R "
          "(L = low y, R = high y)")
    for team in (0, 1):
        counts = "  ".join(f"{ZONE_LABELS[z]}:{zone_entries[team][z]}"
                           for z in ZONE_NAMES)
        print(f"  Team {team + 1}:  {counts}")

    # Write the report to a text file (xG + entries + interval history)
    final_clock = format_clock(frame_count / fps)
    report_path = report_path or os.path.join(
        os.path.dirname(__file__),
        f"match_report_{time.strftime('%Y%m%d_%H%M%S')}.txt")
    try:
        write_report(team_xg, zone_entries, source_desc, final_clock,
                     interval_history, report_path)
        print(f"\n📝 Report saved to: {report_path}")
    except OSError as e:
        print(f"\n⚠️  Could not write report ({e})")

    # Write the same report as an Excel workbook the coach can open directly.
    xlsx_path = os.path.splitext(report_path)[0] + ".xlsx"
    try:
        write_report_xlsx(team_xg, zone_entries, source_desc, final_clock,
                          interval_history, all_shots, xlsx_path)
        print(f"📊 Excel report saved to: {xlsx_path}")
    except ImportError:
        print("\n⚠️  Excel export skipped: openpyxl not installed "
              "(run: pip install openpyxl)")
    except OSError as e:
        print(f"\n⚠️  Could not write Excel report ({e})")

    # Write a per-shot CSV so xG can be re-tuned offline without re-processing
    # the video (filter by dist_m, recompute totals from these rows, etc.).
    shots_path = os.path.splitext(report_path)[0] + "_shots.csv"
    try:
        with open(shots_path, "w", newline="") as f:
            fieldnames = ["clock", "frame", "team", "accepted", "reason", "xg",
                          "speed_mps", "dist_m", "aim_deg", "max_step_m",
                          "x_m", "y_m"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_shots)
        n_acc = sum(s["accepted"] for s in all_shots)
        print(f"📝 Shot log saved to: {shots_path} "
              f"({len(all_shots)} candidates, {n_acc} accepted)")
    except OSError as e:
        print(f"\n⚠️  Could not write shot log ({e})")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Real-time soccer analytics with Expected Goals (xG).")
    parser.add_argument(
        "--video", type=str, default=None,
        help="Path to a video file OR a URL (including YouTube). "
             "If omitted, a live camera is used.")
    parser.add_argument(
        "--report", type=str, default=None,
        help="Path for the end-of-run text report (xG + final-third entries). "
             "Defaults to match_report_<timestamp>.txt in the script folder.")
    parser.add_argument(
        "--stride", type=int, default=1,
        help="Process every Nth frame (skip the rest) for a near-linear "
             "speedup when batch-processing. e.g. --stride 4. Default 1.")
    parser.add_argument(
        "--no-display", action="store_true",
        help="Headless mode: skip all GUI windows and the radar/Voronoi "
             "rendering for maximum throughput. Stats and report still run.")
    parser.add_argument(
        "--flip-attack", action="store_true",
        help="Flip attacking direction (teams switch sides at half-time). "
             "Use this on the second-half run so zone entries are measured "
             "toward the correct goals.")
    parser.add_argument(
        "--swap-teams", action="store_true",
        help="Swap Team 1/Team 2 identities if the auto colour-based labeling "
             "is reversed for your kits (see the kit-colour printout).")
    args = parser.parse_args()
    main(video_path=args.video, report_path=args.report,
         stride=max(1, args.stride), no_display=args.no_display,
         flip_attack=args.flip_attack, swap_teams=args.swap_teams)